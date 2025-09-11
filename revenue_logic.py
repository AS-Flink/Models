# revenue_logic.py (Corrected Version)

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import io
import traceback

# --- MODIFIED IMPORT BLOCK ---
# This new structure provides clearer error messages if a file is missing.
IMPORTS_OK = False
IMPORT_ERROR_MESSAGE = ""
try:
    from imbalance_algorithm_SAP import run_battery_trading as run_battery_trading_SAP
    from self_consumption_PV_PAP import run_battery_trading as run_battery_trading_PAP
    from day_ahead_trading_PAP import run_battery_trading as run_battery_trading_day_ahead
    from imbalance_everything_PAP import run_battery_trading as run_battery_trading_everything_PAP
    IMPORTS_OK = True
except ImportError as e:
    IMPORT_ERROR_MESSAGE = f"Critical Error: Could not import an algorithm file. Please ensure all four trading algorithm scripts (`day_ahead_trading_PAP.py`, `imbalance_algorithm_SAP.py`, etc.) are in the correct directory. Details: {e}"
# --- END MODIFIED BLOCK ---


def run_revenue_model(params, input_df, progress_callback):
    """
    This is the corrected version that replicates the full output logic
    from the original Tkinter application.
    """
    # --- ADD THIS CHECK AT THE TOP ---
    if not IMPORTS_OK:
        return {
            "summary": None,
            "output_file_bytes": None,
            "warnings": [],
            "error": IMPORT_ERROR_MESSAGE
        }
    # --- END CHECK ---

    now = datetime.datetime.now()
    warnings = []
    
    class Cfg: pass
    config = Cfg()
    for k, v in params.items():
        setattr(config, k, v)
    
    # The input dataframe is now passed directly
    config.input_data = input_df

    progress_callback("Starting model run...")

    try:
        # --- 1. Run the selected battery trading algorithm ---
        if params["BATTERY_CONFIG"] == "Onbalanshandel, alleen batterij op SAP":
            df, summary = run_battery_trading_SAP(config, progress_callback=progress_callback)
        elif params["BATTERY_CONFIG"] == "Onbalanshandel, alles op onbalansprijzen":
            df, summary = run_battery_trading_everything_PAP(config, progress_callback=progress_callback)
        elif params["BATTERY_CONFIG"] == "Day-ahead trading, minimaliseer energiekosten":
            df, summary = run_battery_trading_day_ahead(config, progress_callback=progress_callback)
        else: # "Verhogen eigen verbruik PV, alles op day-ahead"
            df, summary = run_battery_trading_PAP(config, progress_callback=progress_callback)
        
        if df is None or not isinstance(df, pd.DataFrame):
            raise ValueError("Model run failed to return a valid DataFrame.")

        progress_callback("Model run complete. Generating Excel output...")
        
        # --- 2. Define the specific columns to export based on the config (REPLICATED LOGIC) ---
        if params["BATTERY_CONFIG"] == "Onbalanshandel, alleen batterij op SAP":
            desired_columns = [
                'regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh',
                'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax',
                'supplier_costs', 'transport_costs', 'total_result_imbalance_SAP'
            ]
        elif params["BATTERY_CONFIG"] == "Onbalanshandel, alles op onbalansprijzen":
            desired_columns = [
                'regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh',
                'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax',
                'supplier_costs', 'transport_costs', 'total_result_imbalance_PAP'
            ]
        elif params["BATTERY_CONFIG"] == "Day-ahead trading, minimaliseer energiekosten":
            desired_columns = [
                'production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2',
                'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs',
                'transport_costs', 'total_result_day_ahead_trading'
            ]
        else:  # For self_consumption_PV_PAP
            desired_columns = [
                'production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2',
                'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs',
                'transport_costs', 'total_result_self_consumption'
            ]

        # Filter the DataFrame to only include columns that exist from the desired list
        df.index.name = 'Datetime'
        existing_columns = [col for col in desired_columns if col in df.columns]
        df_export = df[existing_columns]

        # --- 3. Create a new workbook and populate it correctly (REPLICATED LOGIC) ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import uit Python"

        # Write data using dataframe_to_rows, starting from cell B7
        # dataframe_to_rows includes index and header
        rows = dataframe_to_rows(df_export, index=True, header=True)
        for r_idx, row in enumerate(rows, start=7):
            for c_idx, value in enumerate(row, start=2):
                 ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Create the detailed summary text and merge cell C6:M6
        ws.merge_cells('C6:M6')
        datum_str = now.strftime('%d-%m-%Y %Hu%M')
        optimization_method = summary.get('optimization_method', 'Pyomo optimalisatie')
        summary_text = (
            f"Python run {datum_str}      {params['POWER_MW']} MW      {params['CAPACITY_MWH']} MWh      "
            f"{round(summary.get('total_cycles', 0), 1)} cycli per jaar.      "
            f"Algoritme: {params['BATTERY_CONFIG']}      "
            f"Optimalisatie: {optimization_method}"
        )
        ws['C6'] = summary_text
        
        # Write input parameters to cells W2:W9
        ws['W2'] = params['POWER_MW']
        ws['W3'] = params['CAPACITY_MWH']
        ws['W4'] = params['MIN_SOC']
        ws['W5'] = params['MAX_SOC']
        ws['W6'] = params['EFF_CH']
        ws['W7'] = params['EFF_DIS']
        ws['W8'] = params['SUPPLY_COSTS']
        ws['W9'] = params['TRANSPORT_COSTS']

        # --- 4. Save workbook to in-memory buffer for download ---
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        # Check for any warnings from the model run
        if 'warning_message' in summary and summary['warning_message']:
            warnings.append(summary['warning_message'])
        if 'infeasible_days' in summary and len(summary.get('infeasible_days', [])) > 0:
            warnings.append(f"Model was infeasible for {len(summary['infeasible_days'])} days.")

        progress_callback("Output file generated successfully!")
        
        return {
            "summary": summary,
            "output_file_bytes": output_buffer.getvalue(),
            "warnings": warnings,
            "error": None
        }

    except Exception as e:
        tb = traceback.format_exc()
        print(f"ERROR in revenue_logic: {tb}")
        return {
            "summary": None,
            "output_file_bytes": None,
            "warnings": [],
            "error": f"An error occurred during the model run: {str(e)}"
        }
